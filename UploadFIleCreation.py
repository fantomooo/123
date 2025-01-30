import logging
from openpyxl import load_workbook
from colorama import Fore, Style, init
import sys
import os

# Инициализация colorama
init(autoreset=True)

# Настройка логирования
logging.basicConfig(
    level=logging.INFO,
    format="%(message)s",
    handlers=[logging.StreamHandler(sys.stdout)]
)

def create_upload_files(stockready_file='StockReady.xlsx', output_dir='uploads'):
    """
    Создает текстовые файлы на основе данных из StockReady.xlsx.

    :param stockready_file: Имя исходного Excel-файла.
    :param output_dir: Папка для сохранения текстовых файлов.
    """
    try:
        # Загрузка Excel-файла
        workbook = load_workbook(filename=stockready_file)
        sheet = workbook.active
        logging.info(f"{Fore.GREEN}Loaded '{stockready_file}' successfully.{Style.RESET_ALL}")
    except Exception as e:
        logging.error(f"{Fore.RED}Error loading Excel file '{stockready_file}': {e}{Style.RESET_ALL}")
        return

    # Определение заголовков
    headers = {
        'SKU': 'SKU',
        'ParserLink': 'ParserLink',
        'Stock': 'Stock'
    }

    # Поиск всех колонок, начинающихся с 'SKU'
    sku_columns = []
    max_column = sheet.max_column
    for col in range(1, max_column + 1):
        cell_value = sheet.cell(row=2, column=col).value  # Заголовки находятся во второй строке
        if cell_value == headers['SKU']:
            sku_columns.append(col)

    if not sku_columns:
        logging.error(f"{Fore.RED}No 'SKU' columns found in '{stockready_file}'. Exiting.{Style.RESET_ALL}")
        return

    logging.info(f"{Fore.GREEN}Found {len(sku_columns)} account(s) in '{stockready_file}'.{Style.RESET_ALL}")

    # Создание папки для выходных файлов, если не существует
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)

    # Проход по каждой колонке SKU и обработка соответствующих данных
    for idx, sku_col in enumerate(sku_columns, start=1):
        parser_link_col = sku_col + 1
        stock_col = parser_link_col + 1

        # Проверка, что stock_col не превышает количество колонок
        if stock_col > max_column:
            logging.warning(
                f"{Fore.YELLOW}Missing 'Stock' column for SKU column {sku_col}. Skipping this account.{Style.RESET_ALL}")
            continue

        # Получение названия магазина из первой строки соответствующего столбца SKU
        store_name_cell = f"{sheet.cell(row=1, column=sku_col).coordinate}"
        store_name = sheet.cell(row=1, column=sku_col).value
        if not store_name:
            store_name = f"Store_{idx}"  # Если название не найдено, использовать дефолтное
            logging.warning(f"{Fore.YELLOW}Store name not found in cell {store_name_cell}. Using default name '{store_name}'.{Style.RESET_ALL}")
        else:
            logging.info(f"{Fore.BLUE}Store name for account {idx}: {store_name}{Style.RESET_ALL}")

        # Очистка названия магазина от пробелов для имени файла
        sanitized_store_name = ''.join(store_name.split()) if isinstance(store_name, str) else f"Store_{idx}"

        # Чтение данных из колонок
        data = []
        for row in range(3, sheet.max_row + 1):  # Начинаем с 3-й строки
            sku = sheet.cell(row=row, column=sku_col).value
            stock = sheet.cell(row=row, column=stock_col).value

            # Пропуск строк с пустыми SKU
            if not sku:
                continue

            # Преобразование stock по правилам
            if isinstance(stock, (int, float)):
                quantity = '0' if stock == 0 else '1'
            elif stock is None:
                quantity = ""
            else:
                try:
                    stock_num = int(stock)
                    quantity = '0' if stock_num == 0 else '1'
                except ValueError:
                    quantity = ""

            data.append({
                'SKU': sku,
                'quantity': quantity,
                'handling_time': '3'  # Фиксированное значение
            })

        # Определение имени выходного файла
        output_filename = f"upload_{sanitized_store_name}.txt"
        output_path = os.path.join(output_dir, output_filename)

        # Запись данных в текстовый файл
        try:
            with open(output_path, 'w', encoding='utf-8') as f:
                # Запись заголовка во вторую строку
                header_line = "sku\tprice\tminimum-seller-allowed-price\tmaximum-seller-allowed-price\tquantity\thandling-time\tfulfillment-channel\n"
                f.write(header_line)

                # Запись строк данных начиная с третьей строки
                for entry in data:
                    line = (
                        f"{entry['SKU']}\t"  # SKU
                        f"\t\t\t"             # Пустые поля: price, minimum-seller-allowed-price, maximum-seller-allowed-price
                        f"{entry['quantity']}\t"  # quantity
                        f"{entry['handling_time']}\t"  # handling-time
                        f"\n"  # fulfillment-channel пустой
                    )
                    f.write(line)

            logging.info(f"{Fore.GREEN}Created '{output_filename}' with {len(data)} entries.{Style.RESET_ALL}")
        except Exception as e:
            logging.error(f"{Fore.RED}Error writing to '{output_filename}': {e}{Style.RESET_ALL}")

    logging.info(f"{Fore.CYAN}All upload files have been created in the '{output_dir}' directory.{Style.RESET_ALL}")

if __name__ == "__main__":
    # Путь к файлу StockReady.xlsx и папке для выходных файлов можно задать через аргументы командной строки
    import argparse

    parser = argparse.ArgumentParser(description="Generate upload text files from StockReady.xlsx")
    parser.add_argument('--input', type=str, default='StockReady.xlsx', help='Path to the StockReady.xlsx file')
    parser.add_argument('--output_dir', type=str, default='uploads', help='Directory to save the upload text files')

    args = parser.parse_args()

    create_upload_files(stockready_file=args.input, output_dir=args.output_dir)
