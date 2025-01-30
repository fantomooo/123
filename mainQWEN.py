import requests
import pandas as pd
import concurrent.futures
import re
import sys
import random
import logging
import time
import subprocess  # Добавлено для запуска второго скрипта
from datetime import datetime, timedelta
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry
from lxml import html
from colorama import Fore, Style, init
from openpyxl import load_workbook, Workbook

init(autoreset=True)

# Константы настройки
MAX_THREADS = 40
MAX_RETRIES = 2
TIMEOUT = 5
USER_AGENTS = [
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/119.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/119.0.0.0 Safari/537.36",
    "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/119.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:109.0) Gecko/20100101 Firefox/119.0",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/16.0 Safari/605.1.15",
]
RESULTS_PATTERN = re.compile(r'(\d+(?:,\d+)*)\s+results?\b', re.IGNORECASE)

# Настройка логирования
logging.basicConfig(
    level=logging.INFO,
    format="%(message)s",
    handlers=[logging.StreamHandler(sys.stdout)]
)

# Определение пар столбцов SKU и ParserLink
COLUMN_PAIRS = [
    ('D', 'E'),
    ('G', 'H'),
    ('J', 'K'),
    ('M', 'N'),
    ('P', 'Q'),
    ('S', 'T'),
    ('V', 'W'),
]

# Глобальные переменные для хранения ссылок с ошибками и ошибок прокси
failed_urls = []
proxy_errors_count = {}

def setup_session(proxy=None):
    """
    Настройка сессии с прокси и заголовками.
    """
    session = requests.Session()
    retries = Retry(
        total=MAX_RETRIES,
        backoff_factor=1,
        status_forcelist=[500, 502, 503, 504]
    )
    session.mount('http://', HTTPAdapter(max_retries=retries))
    session.mount('https://', HTTPAdapter(max_retries=retries))
    if proxy:
        session.proxies.update(proxy)
    session.headers.update({
        "User-Agent": random.choice(USER_AGENTS),
        "Accept-Language": "en-US,en;q=0.9",
        "Accept-Encoding": "gzip, deflate, br",
    })
    return session

def load_proxies():
    """
    Загрузка прокси из файла proxies.txt.
    Формат каждой строки: ip:port:user:pwd
    """
    proxies = []
    try:
        with open('proxies.txt', 'r', encoding='utf-8') as f:
            for line in f:
                line = line.strip()
                if line.count(':') == 3:
                    ip, port, user, pwd = line.split(':')
                    proxy_url = f"http://{user}:{pwd}@{ip}:{port}"
                    proxies.append({'http': proxy_url, 'https': proxy_url})
        logging.info(f"{Fore.GREEN}Loaded {len(proxies)} proxies{Style.RESET_ALL}")
    except Exception as e:
        logging.error(f"{Fore.RED}Proxy loading error: {e}{Style.RESET_ALL}")
    return proxies

def check_proxy(proxy):
    """
    Проверка работоспособности прокси на доступ к eBay.
    """
    test_url = "https://www.ebay.com"
    try:
        with setup_session(proxy) as session:
            response = session.get(test_url, timeout=TIMEOUT)
            if response.status_code == 200 and "eBay" in response.text:
                return True
            else:
                return False
    except:
        return False

def process_product(session, url):
    """
    Парсинг страницы по URL и извлечение количества результатов.
    """
    try:
        response = session.get(url, timeout=TIMEOUT)
        response.raise_for_status()
        tree = html.fromstring(response.content)
        result_elements = tree.xpath("//h1[@class='srp-controls__count-heading']")
        if result_elements:
            text = result_elements[0].text_content().strip()
            match = RESULTS_PATTERN.search(text)
            if match:
                return int(match.group(1).replace(',', ''))
        return 0
    except:
        return "Error"

def worker(url, proxy, pair_idx, row, sku, parser_link):
    """
    Рабочая функция для обработки одной ссылки с использованием прокси.
    """
    global failed_urls, proxy_errors_count
    try:
        with setup_session(proxy) as session:
            result = process_product(session, url)
            if result == "Error":
                if proxy and proxy['http'] in proxy_errors_count:
                    proxy_errors_count[proxy['http']] += 1
                elif proxy:
                    proxy_errors_count[proxy['http']] = 1
                # Добавление ссылки в список неудачных
                failed_urls.append({
                    'pair_idx': pair_idx,
                    'row': row,
                    'sku': sku,
                    'parser_link': parser_link,
                    'url': url
                })
            return url, result
    except:
        if proxy and proxy['http'] in proxy_errors_count:
            proxy_errors_count[proxy['http']] += 1
        elif proxy:
            proxy_errors_count[proxy['http']] = 1
        # Добавление ссылки в список неудачных
        failed_urls.append({
            'pair_idx': pair_idx,
            'row': row,
            'sku': sku,
            'parser_link': parser_link,
            'url': url
        })
        return url, "Error"

def save_workbook_with_retries(output_wb, filename='StockReady.xlsx', retries=5, delay=5):
    """
    Попытка сохранить рабочую книгу с повторными попытками при ошибках.
    """
    for attempt in range(1, retries + 1):
        try:
            output_wb.save(filename)
            logging.info(f"{Fore.GREEN}Workbook saved successfully to '{filename}'.{Style.RESET_ALL}")
            return True
        except PermissionError:
            logging.error(f"{Fore.RED}PermissionError: Unable to save '{filename}'. Attempt {attempt} of {retries}. Ensure the file is closed.{Style.RESET_ALL}")
            if attempt < retries:
                logging.info(f"{Fore.YELLOW}Retrying in {delay} seconds...{Style.RESET_ALL}")
                time.sleep(delay)
            else:
                logging.error(f"{Fore.RED}Failed to save '{filename}' after {retries} attempts. Exiting.{Style.RESET_ALL}")
                return False
        except Exception as e:
            logging.error(f"{Fore.RED}Error saving '{filename}': {e}{Style.RESET_ALL}")
            return False
    return False

def main():
    global failed_urls, proxy_errors_count
    error_count = 0
    total_links = 0
    global_processed = 0

    try:
        workbook = load_workbook(filename='Stock_All.xlsx')
        sheet = workbook.active
        logging.info(f"{Fore.GREEN}Loaded 'Stock_All.xlsx' successfully.{Style.RESET_ALL}")
    except Exception as e:
        logging.error(f"{Fore.RED}Error loading Excel file: {e}{Style.RESET_ALL}")
        return

    # Подсчёт общего количества ссылок
    for sku_col, parser_col in COLUMN_PAIRS:
        for row in range(3, sheet.max_row + 1):
            url = sheet[f'{parser_col}{row}'].value
            if url and isinstance(url, str):
                total_links += 1
    logging.info(f"{Fore.GREEN}Total URLs to process: {total_links}{Style.RESET_ALL}")

    proxies = load_proxies()
    valid_proxies = []
    invalid_proxies = []
    if proxies:
        with concurrent.futures.ThreadPoolExecutor(max_workers=MAX_THREADS) as executor:
            future_to_proxy = {executor.submit(check_proxy, proxy): proxy for proxy in proxies}
            for future in concurrent.futures.as_completed(future_to_proxy):
                proxy = future_to_proxy[future]
                try:
                    if future.result():
                        valid_proxies.append(proxy)
                    else:
                        invalid_proxies.append(proxy)
                except Exception as e:
                    invalid_proxies.append(proxy)
    # Упрощённый вывод проверки прокси
    logging.info(f"{Fore.GREEN}All proxies have been checked.{Style.RESET_ALL}")
    logging.info(f"{Fore.GREEN}Valid proxies: {len(valid_proxies)}{Style.RESET_ALL}")
    if invalid_proxies:
        logging.warning(f"{Fore.YELLOW}Invalid proxies: {len(invalid_proxies)}{Style.RESET_ALL}")
        for proxy in invalid_proxies:
            logging.warning(f" - Proxy {proxy['http']}{Style.RESET_ALL}")

    if not valid_proxies and proxies:
        logging.error(f"{Fore.RED}No valid proxies available. Exiting.{Style.RESET_ALL}")
        return

    # Создание нового Excel-файла для результатов
    output_wb = Workbook()
    output_ws = output_wb.active
    output_ws.title = "StockReady"

    # Инициализация столбца для записи результатов
    current_col = 1  # Начинаем с первого столбца

    # Общий прогресс обработки всех ссылок
    overall_start_time = datetime.now()

    for idx, (sku_col, parser_col) in enumerate(COLUMN_PAIRS, start=1):
        urls = []
        for row in range(3, sheet.max_row + 1):
            sku = sheet[f'{sku_col}{row}'].value
            url = sheet[f'{parser_col}{row}'].value
            if url and isinstance(url, str):
                urls.append({'row': row, 'sku': sku, 'url': url, 'parser_link': url})
        logging.info(f"{Fore.GREEN}Loaded {len(urls)} URLs from columns {sku_col}-{parser_col} in 'Stock_All.xlsx'{Style.RESET_ALL}")
        if not urls:
            logging.warning(f"{Fore.YELLOW}No URLs found in columns {sku_col}-{parser_col}. Skipping this pair.{Style.RESET_ALL}")
            # При отсутствии данных для данной пары, оставляем пустные столбцы
            current_col += 3 + 1  # Три столбца данных и один пустой
            continue

        # Получение названия магазина из первой строки соответствующего столбца SKU
        store_name_cell = f"{sku_col}1"
        store_name = sheet[store_name_cell].value
        if not store_name:
            store_name = f"Store_{idx}"  # Если название не найдено, использовать дефолтное
            logging.warning(f"{Fore.YELLOW}Store name not found in cell {store_name_cell}. Using default name '{store_name}'.{Style.RESET_ALL}")
        else:
            logging.info(f"{Fore.BLUE}Store name for pair {idx}: {store_name}{Style.RESET_ALL}")

        results = []
        processed = 0
        total = len(urls)
        start_time = datetime.now()
        if valid_proxies:
            proxy_cycle = iter(valid_proxies)
        else:
            proxy_cycle = iter([None])

        with concurrent.futures.ThreadPoolExecutor(max_workers=MAX_THREADS) as executor:
            futures = []
            for item in urls:
                url = item['url']
                row = item['row']
                sku = item['sku']
                parser_link = item['parser_link']
                try:
                    proxy = next(proxy_cycle)
                except StopIteration:
                    if valid_proxies:
                        proxy_cycle = iter(valid_proxies)
                        proxy = next(proxy_cycle)
                    else:
                        proxy = None
                futures.append(executor.submit(worker, url, proxy, idx, row, sku, parser_link))

            for future in concurrent.futures.as_completed(futures):
                url, count = future.result()
                if count == "Error":
                    error_count += 1
                else:
                    results.append({'url': url, 'stock': count})
                processed += 1
                global_processed += 1

                # Удаление прокси с слишком большим количеством ошибок
                proxies_to_remove = [proxy for proxy, errs in proxy_errors_count.items() if errs >= 5]
                for proxy in proxies_to_remove:
                    logging.warning(f"{Fore.YELLOW}Removing proxy {proxy} due to too many errors{Style.RESET_ALL}")
                    valid_proxies = [p for p in valid_proxies if p['http'] != proxy]
                    del proxy_errors_count[proxy]
                    proxy_cycle = iter(valid_proxies) if valid_proxies else iter([None])

                # Обновление прогресса
                elapsed_time = datetime.now() - start_time
                if processed > 0:
                    avg_time_per_item = elapsed_time / processed
                    remaining_items_pair = total - processed
                    remaining_time_pair = avg_time_per_item * remaining_items_pair
                else:
                    remaining_time_pair = timedelta(seconds=0)

                if global_processed > 0:
                    overall_elapsed_time = datetime.now() - overall_start_time
                    avg_time_per_item_global = overall_elapsed_time / global_processed
                    remaining_items_global = total_links - global_processed
                    remaining_time_global = avg_time_per_item_global * remaining_items_global
                else:
                    remaining_time_global = timedelta(seconds=0)

                # Форматирование времени
                def format_timedelta(td):
                    total_seconds = int(td.total_seconds())
                    hours, remainder = divmod(total_seconds, 3600)
                    minutes, seconds = divmod(remainder, 60)
                    if hours > 0:
                        return f"{hours}h {minutes}m {seconds}s"
                    elif minutes > 0:
                        return f"{minutes}m {seconds}s"
                    else:
                        return f"{seconds}s"

                eta_pair = format_timedelta(remaining_time_pair)
                eta_global = format_timedelta(remaining_time_global)

                percentage_pair = (processed / total) * 100 if total else 100
                percentage_global = (global_processed / total_links) * 100 if total_links else 100

                bar_length = 20
                filled_length_pair = int(bar_length * processed // total) if total else bar_length
                bar_pair = '#' * filled_length_pair + '-' * (bar_length - filled_length_pair)

                filled_length_global = int(bar_length * global_processed // total_links) if total_links else bar_length
                bar_global = '#' * filled_length_global + '-' * (bar_length - filled_length_global)

                # Получение топ-3 прокси с наибольшим количеством ошибок
                top_errors = sorted(proxy_errors_count.items(), key=lambda x: x[1], reverse=True)[:3]
                top_errors_str = ', '.join([f"{proxy}: {errs}" for proxy, errs in top_errors]) if top_errors else "None"

                progress_message = (
                    f"\r{Fore.CYAN}Overall Progress: |{bar_global}| {percentage_global:.2f}% "
                    f"({global_processed}/{total_links}) | Pair {idx}: |{bar_pair}| {percentage_pair:.2f}% "
                    f"({processed}/{total}) | ETA Global: {eta_global} | ETA Pair: {eta_pair} | Errors: {error_count} | Top Proxy Errors: {top_errors_str}{Style.RESET_ALL}"
                )
                sys.stdout.write(progress_message)
                sys.stdout.flush()
        print()  # Для переноса строки после прогресс-бара

        # Подготовка данных для записи в новый Excel
        data = []
        url_to_stock = {res['url']: res['stock'] for res in results}
        for item in urls:
            sku = item['sku']
            url = item['url']
            stock = url_to_stock.get(url, "Error")
            data.append({'SKU': sku, 'ParserLink': url, 'Stock': stock})

        # Запись названия магазина (Store Name) в первую строку, слияние ячеек
        try:
            output_ws.merge_cells(start_row=1, start_column=current_col, end_row=1, end_column=current_col + 2)
            output_ws.cell(row=1, column=current_col, value=store_name)
            logging.info(f"{Fore.BLUE}Inserted store name '{store_name}' in columns {current_col}-{current_col + 2}.{Style.RESET_ALL}")
        except Exception as e:
            logging.error(f"{Fore.RED}Error inserting store name '{store_name}': {e}{Style.RESET_ALL}")

        # Запись заголовков для текущего аккаунта во вторую строку
        output_ws.cell(row=2, column=current_col, value="SKU")
        output_ws.cell(row=2, column=current_col + 1, value="ParserLink")
        output_ws.cell(row=2, column=current_col + 2, value="Stock")

        # Запись данных начиная с третьей строки
        for i, entry in enumerate(data, start=3):
            output_ws.cell(row=i, column=current_col, value=entry['SKU'])
            output_ws.cell(row=i, column=current_col + 1, value=entry['ParserLink'])
            output_ws.cell(row=i, column=current_col + 2, value=entry['Stock'])

        logging.info(f"{Fore.GREEN}Results for pair {idx} ({sku_col}-{parser_col}) written to 'StockReady.xlsx'{Style.RESET_ALL}")

        # Промежуточное сохранение после обработки текущей пары
        if not save_workbook_with_retries(output_wb, 'StockReady.xlsx', retries=5, delay=5):
            logging.error(f"{Fore.RED}Failed to save 'StockReady.xlsx' after multiple attempts. Exiting.{Style.RESET_ALL}")
            sys.exit(1)  # Завершение работы скрипта, так как невозможно сохранить результаты

        current_col += 3 + 1  # Три столбца данных и один пустой

    # Перепроверка ссылок с ошибками
    if failed_urls:
        logging.info(f"{Fore.YELLOW}Reprocessing {len(failed_urls)} failed URLs...{Style.RESET_ALL}")
        reprocess_start_time = datetime.now()
        new_failed_urls = []

        # Создание списка для перепроверки
        reprocess_items = failed_urls.copy()
        failed_urls.clear()  # Очистка списка для возможных новых ошибок

        with concurrent.futures.ThreadPoolExecutor(max_workers=MAX_THREADS) as executor:
            futures = []
            for item in reprocess_items:
                url = item['url']
                pair_idx = item['pair_idx']
                row = item['row']
                sku = item['sku']
                parser_link = item['parser_link']
                # Выбор прокси, исключая те, которые уже имеют >=5 ошибок
                available_proxies = [p for p in valid_proxies if p['http'] not in proxy_errors_count or proxy_errors_count[p['http']] < 5]
                if not available_proxies:
                    available_proxies = [None]
                proxy = random.choice(available_proxies)
                futures.append(executor.submit(worker, url, proxy, pair_idx, row, sku, parser_link))

            for future in concurrent.futures.as_completed(futures):
                url, count = future.result()
                if count == "Error":
                    error_count += 1
                    # Добавление в список новых неудачных
                    for item in reprocess_items:
                        if item['url'] == url:
                            new_failed_urls.append(item)
                            break
                else:
                    # Запись результатов обратно в Excel
                    for item in reprocess_items:
                        if item['url'] == url:
                            pair_idx = item['pair_idx']
                            row = item['row']
                            stock = count
                            # Определение столбцов для записи
                            sku_col, parser_col = COLUMN_PAIRS[pair_idx - 1]
                            # Определение позиции в выходном Excel
                            output_col = (pair_idx - 1) * 4 + 1  # 3 данных + 1 пустой
                            output_ws.cell(row=row, column=output_col + 2, value=stock)  # Столбец "Stock"
                            break

    # Сохранение результатов перепроверки
    if not save_workbook_with_retries(output_wb, 'StockReady.xlsx', retries=5, delay=5):
        logging.error(f"{Fore.RED}Failed to save 'StockReady.xlsx' after reprocessing. Exiting.{Style.RESET_ALL}")
        sys.exit(1)

    if failed_urls:
        logging.warning(f"{Fore.RED}Reprocessing failed for {len(new_failed_urls)} URLs.{Style.RESET_ALL}")
        # Можно повторно попытаться перепроверить их или оставить как ошибки
    else:
        logging.info(f"{Fore.GREEN}All failed URLs have been successfully reprocessed.{Style.RESET_ALL}")

    # Сохранение финального состояния Excel-файла (опционально, так как уже сохранено после каждой пары)
    try:
        output_wb.save('StockReady.xlsx')
        logging.info(f"{Fore.GREEN}All results successfully saved to 'StockReady.xlsx'{Style.RESET_ALL}")
    except PermissionError:
        logging.error(f"{Fore.RED}PermissionError: Unable to save 'StockReady.xlsx'. Please ensure the file is not open in another program.{Style.RESET_ALL}")
    except Exception as e:
        logging.error(f"{Fore.RED}Error saving final 'StockReady.xlsx': {e}{Style.RESET_ALL}")

    # Генерация общей статистики
    try:
        all_results = []
        for pair_idx in range(1, len(COLUMN_PAIRS) + 1):
            # Чтение данных из текущего столбца Stock для статистики
            stock_col = (pair_idx - 1) * 4 + 3  # Каждая пара занимает 4 столбца (3 данных + 1 пустой)
            for row in range(3, output_ws.max_row + 1):  # Начинаем с 3 строки, т.к. 1-ая — название, 2-ая — заголовки
                stock = output_ws.cell(row=row, column=stock_col).value
                if isinstance(stock, int):
                    all_results.append(stock)
        if all_results:
            stats = {}
            for count in all_results:
                stats[count] = stats.get(count, 0) + 1
            logging.info(f"\n{Fore.CYAN}Overall Statistics:{Style.RESET_ALL}")
            for result, cnt in sorted(stats.items()):
                if result == 1:
                    logging.info(f"{Fore.MAGENTA}{cnt} продукт(ов) имеют 1 результат{Style.RESET_ALL}")
                elif result == 0:
                    logging.info(f"{Fore.MAGENTA}{cnt} продукт(ов) имеют 0 результатов{Style.RESET_ALL}")
                else:
                    logging.info(f"{Fore.MAGENTA}{cnt} продукт(ов) имеют {result} результатов{Style.RESET_ALL}")
        else:
            logging.warning(f"{Fore.YELLOW}No successful results to display statistics.{Style.RESET_ALL}")
    except Exception as e:
        logging.error(f"{Fore.RED}Error generating statistics: {e}{Style.RESET_ALL}")

    logging.info(f"\n{Fore.CYAN}Processing completed. Total errors: {error_count}{Style.RESET_ALL}")

    # Добавлено: Ожидание 10 секунд с отсчетом и запуск второго скрипта
    logging.info(f"{Fore.YELLOW}Waiting for 10 seconds before starting uploadFileCreation.py...{Style.RESET_ALL}")
    for remaining in range(10, 0, -1):
        sys.stdout.write(f"\r{Fore.YELLOW}Starting in {remaining} seconds...{Style.RESET_ALL}")
        sys.stdout.flush()
        time.sleep(1)
    print()  # Перенос строки после отсчета
    logging.info(f"{Fore.YELLOW}Launching uploadFileCreation.py...{Style.RESET_ALL}")
    try:
        subprocess.run(["python", "uploadFileCreation.py"], check=True)
        logging.info(f"{Fore.GREEN}uploadFileCreation.py executed successfully.{Style.RESET_ALL}")
    except subprocess.CalledProcessError as e:
        logging.error(f"{Fore.RED}Error executing uploadFileCreation.py: {e}{Style.RESET_ALL}")
    except FileNotFoundError:
        logging.error(f"{Fore.RED}uploadFileCreation.py not found. Please ensure the script exists in the current directory.{Style.RESET_ALL}")

if __name__ == "__main__":
    main()
